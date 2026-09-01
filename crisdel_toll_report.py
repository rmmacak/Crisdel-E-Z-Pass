"""
crisdel_toll_report.py
Core logic for the Crisdel Toll Reconciliation Dashboard.

Parses E-ZPass + SunPass transaction PDFs, cross-references them against
the Crisdel equipment/transponder mapping file, flags high-value
transactions and unmatched transponders/plates, and writes a formatted
Excel workbook.

This module has no Streamlit dependency, so it can also be run standalone
from the command line or imported into other scripts / a notebook.
"""
import json
import os
import re
from datetime import datetime

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets', 'crisdel_logo.png')
ARCHIVE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'reports_archive')
ARCHIVE_INDEX = os.path.join(ARCHIVE_DIR, 'index.json')

# Both E-ZPass and SunPass statements use real drawn grid lines for their
# tables, so telling pdfplumber to trust those lines instead of running its
# default auto-detection heuristics is meaningfully faster with identical
# output (benchmarked: ~15-35% faster, zero row-count difference).
FAST_TABLE_SETTINGS = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}

FONT = "Times New Roman"
FRAUD_THRESHOLD = 10.00

# Crisdel brand colors (pulled from the company logo)
BRAND_NAVY = "163581"
BRAND_LIGHT_BLUE = "C1E1F3"
BRAND_NAVY_TEXT = "0F2560"  # slightly darker navy, for readable body text/titles

COLUMNS = [
    ('Source', 12),
    ('Transaction Type', 16),
    ('Transaction Date', 14),
    ('Transaction Time', 12),
    ('Transponder/Plate (raw)', 20),
    ('Truck #', 10),
    ('License Plate', 14),
    ('Assigned Driver', 18),
    ('Agency', 26),
    ('Location', 34),
    ('Amount', 11),
    ('Match Status', 20),
]

# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------
def norm_plate(p):
    if p is None or isinstance(p, float):
        return None
    p = str(p).strip().upper()
    p = re.sub(r'-[A-Z]{2}$', '', p)     # strip trailing state suffix, e.g. -NJ, -FL
    p = re.sub(r'[^A-Z0-9]', '', p)      # strip spaces/dashes
    return p if p else None


def norm_id(x):
    if x is None or isinstance(x, float):
        return None
    x = str(x).strip()
    x = re.sub(r'[^0-9A-Za-z]', '', x)
    return x if x else None


def parse_amount(s):
    if not s:
        return None
    s = s.strip().replace('$', '').replace(',', '')
    neg = s.startswith('(') and s.endswith(')')
    s = s.strip('()')
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 1. Load Crisdel equipment / transponder mapping
# ---------------------------------------------------------------------------
def load_mapping(path):
    """
    Expects a workbook with a header row on row 2 (row index 1) containing
    at least: Name, Make, Model, License Plate Number, EZPASS Transponder ID,
    Sunpass, Assigned Driver.
    """
    df = pd.read_excel(path, sheet_name=0, header=1)
    df = df.rename(columns={
        'Name': 'Truck #',
        'License Plate Number': 'License Plate',
        'EZPASS Transponder ID': 'EZPass Transponder',
        'Sunpass': 'SunPass Transponder',
        'Assigned Driver': 'Assigned Driver',
    })
    keep = ['Truck #', 'Make', 'Model', 'License Plate', 'EZPass Transponder',
            'SunPass Transponder', 'Assigned Driver', 'On Loan']
    for c in keep:
        if c not in df.columns:
            df[c] = None
    df = df[keep].copy()

    df['_plate_norm'] = df['License Plate'].apply(norm_plate)
    df['_ez_norm'] = df['EZPass Transponder'].apply(norm_id)
    df['_sp_norm'] = df['SunPass Transponder'].apply(norm_id)

    by_plate, by_ez, by_sp = {}, {}, {}
    sp_list = []
    for _, row in df.iterrows():
        rec = row.to_dict()
        if pd.notna(row['_plate_norm']) and row['_plate_norm']:
            by_plate.setdefault(row['_plate_norm'], rec)
        if pd.notna(row['_ez_norm']) and row['_ez_norm']:
            by_ez.setdefault(row['_ez_norm'], rec)
        if pd.notna(row['_sp_norm']) and row['_sp_norm']:
            by_sp.setdefault(row['_sp_norm'], rec)
            sp_list.append((row['_sp_norm'], rec))
    return df, by_plate, by_ez, by_sp, sp_list


def match_vehicle(raw_value, source, by_plate, by_ez, by_sp, sp_list):
    """Return (vehicle_record or None, match_method str)."""
    if not raw_value:
        return None, 'no source id'
    raw = str(raw_value).strip()

    is_platelike = bool(re.search(r'[A-Za-z]', raw)) and not raw.isdigit()

    if is_platelike:
        p = norm_plate(raw)
        if p and p in by_plate:
            return by_plate[p], 'plate'

    idnorm = norm_id(raw)
    if not idnorm:
        return None, 'unparseable'

    if source == 'E-ZPass':
        if idnorm in by_ez:
            return by_ez[idnorm], 'ezpass id'
    elif source == 'SunPass':
        if idnorm in by_sp:
            return by_sp[idnorm], 'sunpass id (exact)'
        # PDF transponder IDs are sometimes a prefix of a longer equipment-list
        # value (extra check digit), or off by leading zeros.
        for sp_norm, rec in sp_list:
            if sp_norm == idnorm:
                return rec, 'sunpass id (exact)'
            if len(idnorm) >= 8 and len(sp_norm) >= 8:
                if sp_norm.startswith(idnorm) or idnorm.startswith(sp_norm):
                    return rec, 'sunpass id (prefix)'
                if sp_norm.lstrip('0') == idnorm.lstrip('0') and idnorm.lstrip('0'):
                    return rec, 'sunpass id (zero-pad)'

    p = norm_plate(raw)
    if p and p in by_plate:
        return p and by_plate[p], 'plate'

    return None, 'unmatched'


# ---------------------------------------------------------------------------
# 2. Parse E-ZPass PDF
# ---------------------------------------------------------------------------
def parse_ezpass(path):
    records = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables(table_settings=FAST_TABLE_SETTINGS):
                if not table:
                    continue
                header = table[0]
                if not header or 'TAG # / PLATE' not in ' '.join([h or '' for h in header]):
                    continue
                for row in table[1:]:
                    if not row or len(row) < 12:
                        continue
                    posted, txn_date, tag_plate, agency, desc, entry_time, entry_plaza, \
                        exit_time, exit_plaza, exit_lane, amount, balance = row[:12]
                    if not tag_plate:
                        continue
                    entry_plaza = (entry_plaza or '').replace('\n', ' ').strip()
                    exit_plaza = (exit_plaza or '').replace('\n', ' ').strip()
                    location = exit_plaza or entry_plaza
                    amt = parse_amount(amount)
                    desc_clean = (desc or '').strip().upper()
                    txn_type = 'Toll' if desc_clean == 'TOLL' else 'Account Payment/Adjustment'
                    records.append({
                        'Source': 'E-ZPass',
                        'Transaction Type': txn_type,
                        'Posted Date': posted,
                        'Transaction Date': txn_date,
                        'Transaction Time': exit_time or entry_time,
                        'Transponder/Plate (raw)': tag_plate.strip(),
                        'Agency': agency,
                        'Location': location,
                        'Amount': abs(amt) if amt is not None else None,
                        'Balance': balance,
                    })
    return records


# ---------------------------------------------------------------------------
# 3. Parse SunPass PDF
# ---------------------------------------------------------------------------
def parse_sunpass(path):
    records = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables(table_settings=FAST_TABLE_SETTINGS):
                if not table:
                    continue
                header = table[0]
                if not header or 'TRANSPONDER' not in ' '.join([h or '' for h in header]):
                    continue
                for row in table[1:]:
                    if not row or len(row) < 12:
                        continue
                    posted, txn_date, txn_time, txn_num, transponder, agency, lane, axle, \
                        desc, debit, credit, balance = row[:12]
                    if not transponder:
                        continue
                    debit_amt = parse_amount(debit)
                    credit_amt = parse_amount(credit)
                    if credit_amt is not None:
                        txn_type = 'Account Payment/Adjustment'
                        amt = abs(credit_amt)
                    else:
                        txn_type = 'Toll'
                        amt = abs(debit_amt) if debit_amt is not None else None
                    records.append({
                        'Source': 'SunPass',
                        'Transaction Type': txn_type,
                        'Posted Date': posted,
                        'Transaction Date': txn_date,
                        'Transaction Time': txn_time,
                        'Transponder/Plate (raw)': transponder.strip(),
                        'Agency': agency,
                        'Location': (desc or '').replace('\n', ' ').strip(),
                        'Amount': amt,
                        'Balance': balance,
                    })
    return records


# ---------------------------------------------------------------------------
# 4. Cross-reference
# ---------------------------------------------------------------------------
def build_combined_dataframe(ezpass_pdf_path, sunpass_pdf_path, mapping_xlsx_path,
                              fraud_threshold=FRAUD_THRESHOLD, progress_cb=None):
    """
    progress_cb: optional callable(str) for status updates (used by the
    Streamlit dashboard to show progress messages).
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    log("Loading Crisdel equipment/transponder mapping...")
    map_df, by_plate, by_ez, by_sp, sp_list = load_mapping(mapping_xlsx_path)

    log("Parsing E-ZPass PDF (this can take a minute for large statements)...")
    ez_records = parse_ezpass(ezpass_pdf_path)

    log("Parsing SunPass PDF (this can take a minute for large statements)...")
    sp_records = parse_sunpass(sunpass_pdf_path)

    log(f"Matching {len(ez_records) + len(sp_records):,} transactions to vehicles...")
    all_records = ez_records + sp_records
    for rec in all_records:
        if rec.get('Transaction Type') != 'Toll':
            rec['Truck #'] = None
            rec['License Plate'] = None
            rec['Assigned Driver'] = None
            rec['Make/Model'] = None
            rec['Match Status'] = 'N/A (Account Payment)'
            rec['Match Method'] = 'not applicable'
            rec['Fraud Flag'] = False
            continue

        vehicle, method = match_vehicle(rec['Transponder/Plate (raw)'], rec['Source'],
                                         by_plate, by_ez, by_sp, sp_list)
        if vehicle:
            rec['Truck #'] = vehicle.get('Truck #')
            rec['License Plate'] = vehicle.get('License Plate')
            rec['Assigned Driver'] = vehicle.get('Assigned Driver')
            rec['Make/Model'] = f"{vehicle.get('Make') or ''} {vehicle.get('Model') or ''}".strip()
            rec['Match Status'] = 'Matched'
            rec['Match Method'] = method
        else:
            rec['Truck #'] = None
            rec['License Plate'] = None
            rec['Assigned Driver'] = None
            rec['Make/Model'] = None
            rec['Match Status'] = 'UNMATCHED'
            rec['Match Method'] = method

        amt = rec.get('Amount')
        rec['Fraud Flag'] = bool(
            rec.get('Transaction Type') == 'Toll'
            and amt is not None
            and amt > fraud_threshold
        )

    df = pd.DataFrame(all_records)
    df['_txn_date_parsed'] = pd.to_datetime(df['Transaction Date'], format='%m/%d/%Y', errors='coerce')
    df = df.sort_values(['_txn_date_parsed', 'Source'], na_position='last').reset_index(drop=True)
    log("Done matching.")
    return df


# ---------------------------------------------------------------------------
# 5. Excel report builder
# ---------------------------------------------------------------------------
header_fill = PatternFill(start_color=BRAND_NAVY, end_color=BRAND_NAVY, fill_type="solid")
header_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
title_font = Font(name=FONT, size=14, bold=True, color=BRAND_NAVY_TEXT)
subtitle_font = Font(name=FONT, size=9, italic=True, color="595959")
fraud_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fraud_font = Font(name=FONT, size=10, color="9C0006")
unmatched_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
payment_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
accent_fill = PatternFill(start_color=BRAND_LIGHT_BLUE, end_color=BRAND_LIGHT_BLUE, fill_type="solid")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
body_font = Font(name=FONT, size=10)
center = Alignment(horizontal="center", vertical="center")


def _add_table(ws, name, header_row, last_data_row, first_col, last_col):
    """Turn a header+data range into a real Excel Table (filter/sort dropdowns).
    Skips gracefully if there are zero data rows, since Excel tables need at
    least one data row below the header."""
    if last_data_row < header_row + 1:
        return
    ref = f"{get_column_letter(first_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"
    table = Table(displayName=name, ref=ref)
    # Row banding/highlighting is handled manually (fraud/unmatched/payment
    # fills), so keep the built-in table style plain to avoid clashing.
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(table)


def _write_title(ws, title, subtitle, ncols, start_row=1):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    ws.cell(row=start_row, column=1, value=title).font = title_font
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=ncols)
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = subtitle_font
    return start_row + 3


def _write_table(ws, data_df, header_row, columns, highlight_fraud=True, highlight_unmatched=True):
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    r = header_row + 1
    for _, row in data_df.iterrows():
        is_fraud = highlight_fraud and bool(row.get('Fraud Flag'))
        is_unmatched = highlight_unmatched and row.get('Match Status') == 'UNMATCHED'
        is_payment = row.get('Match Status') == 'N/A (Account Payment)'
        for i, (name, width) in enumerate(columns, start=1):
            val = row.get(name)
            if name == 'Amount':
                val = float(val) if pd.notna(val) else None
            elif pd.isna(val):
                val = ""  # true blank text so Excel INDEX/MATCH doesn't show 0
            c = ws.cell(row=r, column=i, value=val)
            c.font = fraud_font if is_fraud else body_font
            c.border = border
            if name == 'Amount':
                c.number_format = '$#,##0.00'
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
            if is_fraud:
                c.fill = fraud_fill
            elif is_unmatched:
                c.fill = unmatched_fill
            elif is_payment:
                c.fill = payment_fill
        r += 1
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    return r


def build_workbook(df, fraud_threshold=FRAUD_THRESHOLD, output_path=None, logo_path=DEFAULT_LOGO_PATH):
    wb = openpyxl.Workbook()

    # ---- All Transactions ----
    ws1 = wb.active
    ws1.title = "All Transactions"
    hdr_row = _write_title(
        ws1,
        "Crisdel Toll Reconciliation \u2014 All Transactions",
        f"E-ZPass + SunPass, cross-referenced to Crisdel fleet  |  Generated {datetime.now().strftime('%m/%d/%Y')}  |  "
        f"Red = flagged >${fraud_threshold:.0f}  |  Yellow = unmatched transponder/plate  |  Green = account payment/adjustment",
        len(COLUMNS),
    )
    next_row = _write_table(ws1, df, hdr_row, COLUMNS)
    last_data_row = next_row - 1
    data_first, data_last = hdr_row + 1, last_data_row
    _add_table(ws1, "AllTransactionsTable", hdr_row, last_data_row, 1, len(COLUMNS))

    # ---- Flagged - Over $X ----
    ws2 = wb.create_sheet(f"Flagged - Over ${fraud_threshold:.0f}")
    fraud_df = df[df['Fraud Flag'] == True].copy()
    hdr_row2 = _write_title(
        ws2,
        f"Flagged Transactions \u2014 Over ${fraud_threshold:.0f} (Potential Fraud Review)",
        f"{len(fraud_df)} transactions  |  Generated {datetime.now().strftime('%m/%d/%Y')}",
        len(COLUMNS),
    )
    fraud_next_row = _write_table(ws2, fraud_df, hdr_row2, COLUMNS, highlight_fraud=True, highlight_unmatched=False)
    _add_table(ws2, "FlaggedTable", hdr_row2, fraud_next_row - 1, 1, len(COLUMNS))

    # ---- Unmatched ----
    ws3 = wb.create_sheet("Unmatched Plates-Transponders")
    unmatched_df = df[df['Match Status'] == 'UNMATCHED'].copy()
    hdr_row3 = _write_title(
        ws3,
        "Unmatched Transponders / Plates",
        f"{len(unmatched_df)} transactions could not be matched to a Crisdel vehicle in the mapping file  |  "
        f"Generated {datetime.now().strftime('%m/%d/%Y')}",
        len(COLUMNS),
    )
    unmatched_next_row = _write_table(ws3, unmatched_df, hdr_row3, COLUMNS, highlight_fraud=False, highlight_unmatched=True)
    _add_table(ws3, "UnmatchedTable", hdr_row3, unmatched_next_row - 1, 1, len(COLUMNS))

    # ---- Summary by Truck (formula-driven) ----
    trucks = sorted(df['Truck #'].dropna().unique().tolist())
    ws4 = wb.create_sheet("Summary by Truck", 1)
    SHEET = "'All Transactions'"
    # Column letters in "All Transactions" after removing the Posted Date column:
    # A=Source B=Transaction Type C=Transaction Date D=Transaction Time
    # E=Transponder F=Truck# G=License Plate H=Assigned Driver I=Agency J=Location K=Amount L=Match Status
    COL = {'Truck #': 'F', 'License Plate': 'G', 'Assigned Driver': 'H', 'Amount': 'K'}

    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws4.cell(row=1, column=1, value="Crisdel Toll Reconciliation \u2014 Summary by Truck").font = title_font
    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws4.cell(row=2, column=1,
             value=f"{len(trucks)} vehicles with matched transactions this period  |  "
                   f"Generated {datetime.now().strftime('%m/%d/%Y')}").font = subtitle_font

    cols4 = [("Truck #", 12), ("Assigned Driver", 20), ("License Plate", 14),
             ("Transaction Count", 16), ("Total Spend", 14), (f"Flagged Txns (>${fraud_threshold:.0f})", 18)]
    hdr_row_s = 4
    for i, (name, width) in enumerate(cols4, start=1):
        cc = ws4.cell(row=hdr_row_s, column=i, value=name)
        cc.font, cc.fill, cc.alignment, cc.border = header_font, header_fill, center, border
        ws4.column_dimensions[get_column_letter(i)].width = width

    r = hdr_row_s + 1
    for truck in trucks:
        ws4.cell(row=r, column=1, value=truck)
        ws4.cell(row=r, column=2,
                 value=f'=IFERROR(INDEX({SHEET}!${COL["Assigned Driver"]}${data_first}:${COL["Assigned Driver"]}${data_last},'
                       f'MATCH($A{r},{SHEET}!${COL["Truck #"]}${data_first}:${COL["Truck #"]}${data_last},0))&"","")')
        ws4.cell(row=r, column=3,
                 value=f'=IFERROR(INDEX({SHEET}!${COL["License Plate"]}${data_first}:${COL["License Plate"]}${data_last},'
                       f'MATCH($A{r},{SHEET}!${COL["Truck #"]}${data_first}:${COL["Truck #"]}${data_last},0))&"","")')
        ws4.cell(row=r, column=4,
                 value=f'=COUNTIF({SHEET}!${COL["Truck #"]}${data_first}:${COL["Truck #"]}${data_last},$A{r})')
        ws4.cell(row=r, column=5,
                 value=f'=SUMIFS({SHEET}!${COL["Amount"]}${data_first}:${COL["Amount"]}${data_last},'
                       f'{SHEET}!${COL["Truck #"]}${data_first}:${COL["Truck #"]}${data_last},$A{r})')
        ws4.cell(row=r, column=6,
                 value=f'=COUNTIFS({SHEET}!${COL["Truck #"]}${data_first}:${COL["Truck #"]}${data_last},$A{r},'
                       f'{SHEET}!${COL["Amount"]}${data_first}:${COL["Amount"]}${data_last},">{fraud_threshold}")')
        for col_idx in range(1, len(cols4) + 1):
            cell = ws4.cell(row=r, column=col_idx)
            cell.font, cell.border = body_font, border
            if col_idx == 5:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    _add_table(ws4, "SummaryByTruckTable", hdr_row_s, r - 1, 1, len(cols4))

    total_row = r
    ws4.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
    ws4.cell(row=total_row, column=4, value=f'=SUM(D{hdr_row_s+1}:D{total_row-1})').font = Font(name=FONT, size=10, bold=True)
    tc = ws4.cell(row=total_row, column=5, value=f'=SUM(E{hdr_row_s+1}:E{total_row-1})')
    tc.font = Font(name=FONT, size=10, bold=True)
    tc.number_format = '$#,##0.00'
    ws4.cell(row=total_row, column=6, value=f'=SUM(F{hdr_row_s+1}:F{total_row-1})').font = Font(name=FONT, size=10, bold=True)
    for col_idx in range(1, len(cols4) + 1):
        ws4.cell(row=total_row, column=col_idx).border = border
    ws4.freeze_panes = ws4.cell(row=hdr_row_s + 1, column=1).coordinate

    # ---- README ----
    cover = wb.create_sheet("README", 0)
    cover.sheet_view.showGridLines = False

    title_row = 1
    if logo_path and os.path.exists(logo_path):
        logo_img = XLImage(logo_path)
        logo_img.width = 100
        logo_img.height = 69
        cover.add_image(logo_img, 'A1')
        title_row = 6  # clear of the logo, which spans roughly rows 1-4 at this size

    cover.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=6)
    cover.cell(row=title_row, column=1, value="Crisdel Toll Reconciliation Report").font = \
        Font(name=FONT, size=16, bold=True, color=BRAND_NAVY_TEXT)
    cover.merge_cells(start_row=title_row + 1, start_column=1, end_row=title_row + 1, end_column=6)
    cover.cell(row=title_row + 1, column=1,
               value="E-ZPass + SunPass, monthly, cross-referenced to Crisdel fleet").font = subtitle_font

    toll_df = df[df['Transaction Type'] == 'Toll']
    lines = [
        ("Tabs in this workbook:", ""),
        ("  All Transactions", "Every toll transaction, with Truck #, Plate, and Driver matched in."),
        ("  Summary by Truck", "Transaction count, total spend, and flagged count per vehicle (live formulas)."),
        (f"  Flagged - Over ${fraud_threshold:.0f}", "Every transaction above the threshold \u2014 review for potential fraud."),
        ("  Unmatched Plates-Transponders", "Transponders/plates with no match in the Crisdel mapping file."),
        ("", ""),
        ("Color key (All Transactions tab):", ""),
        ("  Red row", f"Flagged \u2014 transaction over ${fraud_threshold:.0f}"),
        ("  Yellow row", "Unmatched \u2014 no corresponding vehicle in the mapping file"),
        ("  Green row", "Account payment / balance replenishment (not a toll charge)"),
        ("", ""),
        ("This period's totals:", ""),
        ("  Total toll transactions", f"{len(toll_df):,}"),
        ("  Total toll spend", f"${toll_df['Amount'].sum():,.2f}"),
        ("  Matched to a Crisdel vehicle", f"{(toll_df['Match Status']=='Matched').sum():,}"),
        ("  Unmatched", f"{(toll_df['Match Status']=='UNMATCHED').sum():,}"),
        (f"  Flagged (over ${fraud_threshold:.0f})", f"{toll_df['Fraud Flag'].sum():,}"),
    ]
    r = title_row + 3
    for label, val in lines:
        is_header = val == "" and label and not label.startswith(" ")
        cover.cell(row=r, column=1, value=label).font = (
            Font(name=FONT, size=11, bold=True, color=BRAND_NAVY_TEXT) if is_header else body_font)
        vc = cover.cell(row=r, column=2, value=val)
        vc.font = Font(name=FONT, size=10, bold=True) if not is_header and val else body_font
        r += 1
    cover.column_dimensions['A'].width = 32
    cover.column_dimensions['B'].width = 65

    wb['All Transactions'].sheet_properties.tabColor = BRAND_NAVY
    wb['Summary by Truck'].sheet_properties.tabColor = BRAND_LIGHT_BLUE
    wb[f"Flagged - Over ${fraud_threshold:.0f}"].sheet_properties.tabColor = "C00000"
    wb['Unmatched Plates-Transponders'].sheet_properties.tabColor = "BF8F00"
    wb['README'].sheet_properties.tabColor = BRAND_NAVY

    if output_path:
        wb.save(output_path)
    return wb


# ---------------------------------------------------------------------------
# 6. Report history / archive
# ---------------------------------------------------------------------------
def _load_archive_index():
    if not os.path.exists(ARCHIVE_INDEX):
        return []
    try:
        with open(ARCHIVE_INDEX, 'r') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return []


def _save_archive_index(entries):
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    with open(ARCHIVE_INDEX, 'w') as f:
        json.dump(entries, f, indent=2)


def save_report_to_archive(wb, df, fraud_threshold=FRAUD_THRESHOLD):
    """Save a generated workbook into the persistent archive and record its
    stats in a small JSON index, so Report History can list past reports
    without having to re-open every .xlsx file."""
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    ts = datetime.now()
    fname = f"Crisdel_Toll_Report_{ts.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    fpath = os.path.join(ARCHIVE_DIR, fname)
    wb.save(fpath)

    toll_df = df[df['Transaction Type'] == 'Toll']
    entry = {
        'filename': fname,
        'generated_at': ts.isoformat(),
        'label': ts.strftime('%B %Y \u2014 generated %m/%d/%Y %I:%M %p'),
        'toll_transactions': int(len(toll_df)),
        'total_spend': float(toll_df['Amount'].sum()) if len(toll_df) else 0.0,
        'matched': int((toll_df['Match Status'] == 'Matched').sum()),
        'unmatched': int((toll_df['Match Status'] == 'UNMATCHED').sum()),
        'flagged': int(toll_df['Fraud Flag'].sum()),
        'fraud_threshold': fraud_threshold,
    }
    entries = _load_archive_index()
    entries.append(entry)
    _save_archive_index(entries)
    return entry


def list_archived_reports():
    """Newest-first list of archived report metadata dicts."""
    entries = _load_archive_index()
    return sorted(entries, key=lambda e: e['generated_at'], reverse=True)


def get_archived_report_path(filename):
    """Resolve an archived report's filename to a safe path on disk, or
    None if it doesn't exist. Guards against path traversal since the
    filename ultimately comes from user-facing UI state."""
    safe_name = os.path.basename(filename)
    path = os.path.join(ARCHIVE_DIR, safe_name)
    return path if os.path.exists(path) else None
